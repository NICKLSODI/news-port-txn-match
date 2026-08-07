# -*- coding: utf-8 -*-
"""ดึงตารางข้อมูลอ้างอิงออกจากไฟล์ spec (STEP2/3/4) มาเป็น JSON

ทำครั้งเดียวตอน setup แล้ว runtime อ่าน JSON — spec เป็นแหล่งความจริงเดียว
ไม่ hardcode รายการไว้ในโค้ด (หลักการข้อ 5 / R#23)

    python -m scripts.build_refdata

ผลลัพธ์: backend/app/refdata/generated.json
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SPEC = ROOT / "memie"
DATA = ROOT / "data"
OUT = Path(__file__).resolve().parents[1] / "app" / "refdata" / "generated.json"

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app.tables import DR_NAME_ALIAS, DR_ISSUER_SUFFIXES  # noqa: E402
from app import news as _news                              # noqa: E402

STEP2 = SPEC / "STEP2 - Content Inventory & Cadence v2.4.xlsx"
STEP3 = SPEC / "STEP3 - News Tagging Schema v1.8.xlsx"
STEP4 = SPEC / "STEP4 - Instrument Mapping v1.7.xlsx"

TICKER_RE = re.compile(r"^[A-Z][A-Z0-9&.\-]*$")


def rows(path: Path, sheet: str) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    out = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return out


def s(v) -> str:
    return "" if v is None else str(v).strip()


def after(table: list[list], marker: str) -> list[list]:
    """คืนแถวที่อยู่หลังแถวซึ่งมีคำว่า marker"""
    for i, r in enumerate(table):
        if any(marker in s(c) for c in r):
            return table[i + 1:]
    raise KeyError(f"marker not found: {marker!r}")


def num(v):
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


# ==========================================================================

def thai_sector() -> dict:
    """B3 — Thai Stock Coverage List 97 หุ้น (sector / rating / esg / target price)"""
    out = {}
    for r in after(rows(STEP4, "B3 sector หุ้นไทย"), "รายการทั้งหมด"):
        tic, sec = s(r[0]), s(r[1])
        if not TICKER_RE.match(tic) or not sec or sec == "-":
            continue
        out[tic] = {
            "sector": sec,
            "rating": s(r[2]) or None,
            "esg": (s(r[3]) or None) if s(r[3]) != "-" else None,
            "last_close": num(r[4]),
            "target_price": num(r[5]),
        }
    return out


def dr_parent() -> tuple[dict, list]:
    """B1 + R4.24-R4.27 — DR เลขท้าย 2 หลัก -> หุ้นแม่

    คืน (map ที่ใช้ได้, รายการที่ยังแปลงไม่ได้)
    ทุกแถวเป็นระดับ inferred (R4.28 ต้องติดป้าย DR ให้ RM เห็น)
    """
    resolved, pending = {}, []
    for r in after(rows(STEP4, "B1 DR เลข80"), "รายการ DR ทั้งหมด"):
        code, name, parent = s(r[0]), s(r[1]), s(r[2])
        if not code or not TICKER_RE.match(code):
            continue
        if parent and parent != "-":
            resolved[code] = {"parent": parent, "source": "spec_b1", "confidence": "inferred"}
        elif name and name.upper() in DR_NAME_ALIAS:
            resolved[code] = {"parent": DR_NAME_ALIAS[name.upper()],
                              "source": "alias_table", "confidence": "inferred"}
        else:
            pending.append({"code": code, "name": name,
                            "reason": "ไม่มีในตารางชื่อพ้อง (R4.27)"})
    return resolved, pending


def bond_issuer() -> dict:
    """R4.10 — หุ้นกู้ไทย -> บริษัทผู้ออก"""
    out = {}
    for r in after(rows(STEP4, "หุ้นกู้เป็นผู้ออก"), "หุ้นกู้ทั้งหมดที่ลูกค้าถือ"):
        code, issuer = s(r[0]), s(r[1])
        if code and issuer and TICKER_RE.match(code) and TICKER_RE.match(issuer):
            out[code] = issuer
    return out


def tfex_underlying() -> dict:
    """A4 — underlying ของสัญญา TFEX 44 ตัว (หุ้นรายตัว vs ดัชนี/สินค้า)"""
    out = {}
    for r in after(rows(STEP4, "A4 TFEX"), "underlying ทั้งหมดที่พบ"):
        u, name, kind = s(r[0]), s(r[1]), s(r[2])
        if u and kind:
            out[u] = {"name": name, "kind": "single_stock" if "หุ้นรายตัว" in kind else "index"}
    return out


def fund_underlying() -> tuple[dict, dict]:
    """B2 — กองทุน -> underlying (Inferred) + ตาราง keyword -> underlying"""
    funds, kw = {}, {}
    for r in after(rows(STEP4, "B2 กองทุน"), "รายการกองทุนทั้งหมด"):
        code, under, keyword, level = s(r[0]), s(r[1]), s(r[2]), s(r[3])
        if not code:
            continue
        funds[code] = {"underlying": under or None, "keyword": keyword or None,
                       "level": level or "Unknown"}
        if keyword and under:
            kw.setdefault(keyword, under)
    return funds, kw


def entity_dictionary() -> dict:
    """C1 — รหัสสินทรัพย์ -> ชื่อเรียกหลายแบบ (R4.42-R4.47)"""
    out = {}
    for r in after(rows(STEP4, "C1 Entity Dictionary"), "seed dictionary"):
        code, aliases = s(r[0]), s(r[1])
        if not code or not aliases:
            continue
        names = [a.strip() for a in aliases.split(",") if a.strip()]
        out[code] = {"aliases": names, "customers": int(num(r[2]) or 0),
                     "status": s(r[3])}
    return out


def related_groups() -> list:
    """C2 — กลุ่มหุ้นที่ระบบค้นพบเองจากการถูกกล่าวถึงในบทความเดียวกัน (R3.29-R3.36)"""
    out = []
    for r in after(rows(STEP4, "C2 ความสัมพันธ์หุ้น"), "กลุ่มที่ระบบค้นพบเอง"):
        name, members = s(r[0]), s(r[1])
        if not name or not members or name.startswith("ข้อสังเกต"):
            if name.startswith("ข้อสังเกต"):
                break
            continue
        toks = [re.sub(r"\s*\(.*?\)", "", m).strip() for m in members.split(",")]
        toks = [t for t in toks if t]
        if len(toks) >= 2:
            out.append({"group": name, "members": toks})
    return out


def macro_topics() -> tuple[dict, list]:
    """R3.37-R3.40 — คำค้น macro ที่ทดสอบผ่าน + คำที่ห้ามใช้"""
    table = rows(STEP3, "กฎ macro_topic")
    allowed, banned = {}, []
    body = after(table, "คำค้นที่ทดสอบกับ")
    for r in body:
        topic, kws = s(r[0]), s(r[1])
        if "ห้ามใช้เด็ดขาด" in topic:
            break
        if topic and kws and kws != "-":
            allowed[topic] = [k.strip() for k in re.split(r"·|,", kws) if k.strip()]
    for r in after(table, "ห้ามใช้เด็ดขาด"):
        word = s(r[0])
        if not word or "บันทึกการตัดสินใจ" in word:
            break
        banned.append(word)
    return allowed, banned


def content_inventory() -> list:
    """STEP2 — บัญชี 51 หมวด พร้อมความถี่ / มี ticker / product_type / ตัดสิน"""
    out = []
    table = rows(STEP2, "บัญชีเนื้อหา")
    for r in table[3:]:
        cat, sub, name = s(r[0]), s(r[1]), s(r[2])
        if not sub or not cat or sub == "Subcategory":
            continue
        for one in [x.strip() for x in sub.split("/")]:
            if not one:
                continue
            out.append({
                "category": cat.rstrip("-") or cat,
                "category_raw": cat,
                "subcategory": one,
                "name": name,
                "total": num(r[3]),
                "per_week": num(r[4]),
                "has_ticker_pct": s(r[5]),
                "product_type": s(r[6]),
                "role": "reference" if s(r[7]).lower() == "reference" else "content",
                "decision": s(r[8]),
            })
    return out


def dr_names_step3() -> dict:
    """STEP3 ชีต DR กับหุ้นแม่ — รหัส DR -> ชื่อที่ถอดได้ (R3.26)"""
    out = {}
    for r in after(rows(STEP3, "DR กับหุ้นแม่"), "รายการ DR ที่ลูกค้าถือจริง"):
        code, name = s(r[0]), s(r[1])
        if code and name and TICKER_RE.match(code):
            out[code] = name
    return out


def portfolio_codes() -> set[str]:
    """รหัสสินทรัพย์ที่ลูกค้าถือจริง — ใช้ตรวจ alias DR ตาม R3.24

    คืนทั้งรหัสดิบและรหัสกลาง เพราะ alias ในตารางเขียนเป็นรหัสกลาง (FPT:xstc)
    แต่ไฟล์พอร์ตเก็บดิบตามที่ต้นทางส่งมา ซึ่ง MIC ปนตัวพิมพ์ใหญ่อยู่ 407 รหัส
    (R1.2) ถ้าเทียบตรงตัวจะทิ้ง alias ที่ถูกต้องทิ้งไปเงียบ ๆ
    """
    f = next((p for p in DATA.glob("*.xlsx") if "ortfolio" in p.name), None)
    if not f:
        return set()
    wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    ws = wb["result"]
    it = ws.iter_rows(values_only=True)
    hdr = [s(c) for c in next(it)]
    i = hdr.index("product_code")
    raw = {s(r[i]) for r in it if s(r[i])}
    wb.close()

    from app.mapping import offshore_id                     # noqa: PLC0415

    out = set(raw)
    for code in raw:
        if ":" in code:
            tic, mic = code.rsplit(":", 1)
            out.add(offshore_id(tic, mic))                  # FPT:XSTC -> FPT:xstc
    return out


# ==========================================================================

def thai_sector_refreshed() -> dict:
    """B3 จาก Excel (แช่แข็งตอนส่งมอบสเปค) + เติมทับด้วย PDF Recommendation ล่าสุดจาก
    Cafe Invest ถ้าดึงได้ (สดกว่า อัปเดตทุกสัปดาห์) — Excel เป็นฐานที่ใช้ได้เสมอแม้ออฟไลน์
    ดึงสดไม่สำเร็จก็ไม่เป็นไร ใช้ Excel ต่อไปเฉย ๆ ไม่ error (R2.8)

    เติมทับเท่านั้น ไม่ลบ — หุ้นที่ Excel มีแต่ PDF รอบนี้ไม่โผล่ (เช่นหลุด coverage ชั่วคราว)
    ยังอยู่ในผลลัพธ์ ไม่หายไปเงียบ ๆ
    """
    base = thai_sector()
    excel_count = len(base)
    live, problem = _news.thai_coverage_live()
    if problem:
        print(f"  Thai Coverage สด: ดึงไม่สำเร็จ ({problem}) — ใช้ Excel STEP4 ต่อไป")
        return base
    added = [t for t in live if t not in base]
    changed = [t for t in live if t in base and (base[t]["rating"] != live[t]["rating"]
                                                 or base[t]["sector"] != live[t]["sector"])]
    base.update(live)
    print(f"  Thai Coverage สด: ดึงได้ {len(live)} ตัว "
          f"(Excel มี {excel_count} ตัว) — ใหม่ {len(added)} · rating/sector เปลี่ยน {len(changed)}")
    if added:
        print(f"    ใหม่: {', '.join(sorted(added))}")
    if changed:
        print(f"    เปลี่ยน: {', '.join(sorted(changed))}")
    return base


def main() -> None:
    print("reading spec...")
    dr_map, dr_pending = dr_parent()
    funds, fund_kw = fund_underlying()
    macro_allowed, macro_banned = macro_topics()
    held = portfolio_codes()

    # R3.24 / R4.35 — alias ที่ชี้ไปรหัสที่ไม่มีใครถือ ไม่ใช้ ให้เข้ารายงานแทน
    dropped = []
    if held:
        for code, info in list(dr_map.items()):
            if info["source"] == "alias_table" and info["parent"] not in held:
                dropped.append({"code": code, "parent": info["parent"],
                                "reason": "รหัสหุ้นแม่ที่ alias ชี้ไป ไม่พบในพอร์ตลูกค้า"})
                dr_map.pop(code)

    out = {
        "meta": {
            "source_files": ["STEP2 v2.4", "STEP3 v1.8", "STEP4 v1.7"],
            "note": "generated by scripts/build_refdata.py — do not edit by hand",
        },
        "thai_sector": thai_sector_refreshed(),
        "dr_parent": dr_map,
        "dr_pending": dr_pending + dropped,
        "dr_names": dr_names_step3(),
        "bond_issuer": bond_issuer(),
        "tfex_underlying": tfex_underlying(),
        "fund_underlying": funds,
        "fund_keywords": fund_kw,
        "entity_dictionary": entity_dictionary(),
        "related_groups": related_groups(),
        "macro_keywords": macro_allowed,
        "macro_banned": macro_banned,
        "content_inventory": content_inventory(),
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"wrote {OUT}")
    for k, v in out.items():
        if k == "meta":
            continue
        print(f"  {k:22s} {len(v)}")
    sectors = {v["sector"] for v in out["thai_sector"].values()}
    print(f"  -> {len(sectors)} sectors, {len(out['dr_parent'])} DR resolved, "
          f"{len(out['dr_pending'])} DR pending review")


if __name__ == "__main__":
    main()
